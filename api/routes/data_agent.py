"""Built-in Data Agent routes for upload, verification, anchoring, and reuse."""

from __future__ import annotations

import csv
import contextlib
import hashlib
import io
import json
import os
import re
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import FileResponse
from pydantic import BaseModel, ConfigDict, Field, ValidationError
from sqlalchemy.orm import Session

from shared.database import DataAsset, get_db
from shared.hedera.client import get_hedera_client
from shared.a2a.models import AgentCapability, AgentCard, MessagePayload, MessageResponse
from shared.a2a.protocol import (
    build_agent_card_payload,
    build_completed_task_response,
    build_error_response,
    extract_message_text_and_metadata,
)
from shared.hol_client import (
    HOL_DIRECT_SESSION_PREFIX,
    HolAgentSummary,
    HolClientError,
    check_sidecar_health as hol_check_sidecar_health,
    coerce_hol_broker_response,
    create_session as hol_create_session,
    search_agents as hol_search_agents,
    send_message as hol_send_message,
    should_use_direct_chat_fallback,
)
from shared.hol_agent_usability import (
    apply_hol_agent_usability,
    get_hol_agent_verification_map,
    is_hol_hard_availability_failure,
    record_hol_agent_hard_failure,
    record_hol_agent_success,
)

router = APIRouter()

MAX_UPLOAD_BYTES = 25 * 1024 * 1024
ALLOWED_EXTENSIONS = {".csv", ".tsv", ".json", ".txt", ".xlsx", ".zip"}
ALLOWED_CLASSIFICATIONS = {"failed", "underused"}
ALLOWED_VISIBILITY = {"private", "org", "public"}
DEFAULT_STORAGE_DIR = Path(__file__).resolve().parents[2] / "data" / "data_agent_uploads"
PINATA_PIN_JSON_URL = "https://api.pinata.cloud/pinning/pinJSONToIPFS"
DEFAULT_HOL_DATA_CAPABILITIES = ["data", "analysis", "dataset"]


class SimilarDataset(BaseModel):
    """Compact representation for related datasets."""

    id: str
    title: str
    data_classification: str
    tags: List[str]
    similarity_score: float


class DataAssetResponse(BaseModel):
    """Serialized dataset metadata for API responses."""

    model_config = ConfigDict(from_attributes=True)

    id: str
    title: str
    description: Optional[str] = None
    lab_name: str
    uploader_name: Optional[str] = None
    data_classification: str
    tags: List[str]
    intended_visibility: str
    filename: str
    size_bytes: int
    content_type: Optional[str] = None
    sha256: str
    created_at: str
    verification_status: str
    proof_status: str
    manifest_cid: Optional[str] = None
    reuse_count: int
    last_reused_at: Optional[str] = None
    failed_reason: Optional[str] = None
    reuse_domains: List[str]


class DataAssetDetailResponse(DataAssetResponse):
    """Detailed dataset response with trust and similarity context."""

    verification_report: Optional[Dict[str, Any]] = None
    proof_bundle: Optional[Dict[str, Any]] = None
    similar_datasets: List[SimilarDataset] = Field(default_factory=list)
    hol_sessions: List[Dict[str, Any]] = Field(default_factory=list)


class DataAssetListResponse(BaseModel):
    """Paginated datasets listing."""

    total: int
    limit: int
    offset: int
    datasets: List[DataAssetResponse]


class UploadDatasetResponse(DataAssetResponse):
    """Upload response payload."""

    message: str = Field(default="Dataset uploaded successfully.")


class DatasetProofResponse(BaseModel):
    """Proof bundle payload for judges and downstream tooling."""

    dataset_id: str
    file_sha256: str
    manifest_cid: Optional[str] = None
    manifest_sha256: Optional[str] = None
    manifest_gateway_url: Optional[str] = None
    hcs_topic_id: Optional[str] = None
    hcs_message_status: Optional[str] = None
    anchor_payload: Optional[Dict[str, Any]] = None
    anchored_at: Optional[str] = None
    verification_status: str
    verification_report: Optional[Dict[str, Any]] = None
    proof_status: str


class DatasetCitationResponse(BaseModel):
    """JSON citation payload for dataset reuse."""

    citation: Dict[str, Any]


class ReuseEventResponse(BaseModel):
    """Response for manual dataset reuse registration."""

    dataset_id: str
    reuse_count: int
    last_reused_at: str
    message: str


class DatasetHolUseRequest(BaseModel):
    """Request payload for using a HOL data agent from a dataset."""

    uaid: Optional[str] = None
    search_query: Optional[str] = None
    required_capabilities: Optional[List[str]] = None
    instructions: Optional[str] = None
    transport: Optional[str] = None
    as_uaid: Optional[str] = None
    limit: int = Field(default=10, ge=1, le=25)


class DatasetHolUseResponse(BaseModel):
    """Response payload for a HOL data-agent interaction."""

    success: bool
    selected_agent: Dict[str, Any]
    session_id: str
    broker_response: Dict[str, Any] = Field(default_factory=dict)


class DatasetHolUseErrorDetail(BaseModel):
    """Structured HOL discovery diagnostics for UI/debug surfaces."""

    message: str
    search_queries: List[str] = Field(default_factory=list)
    discovered_candidates: List[Dict[str, Any]] = Field(default_factory=list)
    rejected_candidates: List[Dict[str, Any]] = Field(default_factory=list)
    attempted_errors: List[str] = Field(default_factory=list)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _get_storage_dir() -> Path:
    configured = os.getenv("DATA_AGENT_STORAGE_DIR")
    if configured:
        return Path(configured).expanduser().resolve()
    return DEFAULT_STORAGE_DIR


def _parse_string_list(raw_value: Optional[str], field_name: str) -> List[str]:
    if not raw_value:
        return []

    value = raw_value.strip()
    if not value:
        return []

    if value.startswith("["):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"{field_name} must be comma-separated text or a JSON array",
            ) from exc
        if not isinstance(parsed, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"{field_name} JSON must be an array of strings",
            )
        return [str(item).strip() for item in parsed if str(item).strip()]

    return [item.strip() for item in value.split(",") if item.strip()]


def _validate_classification(value: str) -> str:
    normalized = (value or "").strip().lower()
    if normalized not in ALLOWED_CLASSIFICATIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"data_classification must be one of {sorted(ALLOWED_CLASSIFICATIONS)}",
        )
    return normalized


def _validate_visibility(value: str) -> str:
    normalized = (value or "private").strip().lower()
    if normalized not in ALLOWED_VISIBILITY:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"intended_visibility must be one of {sorted(ALLOWED_VISIBILITY)}",
        )
    return normalized


def _default_meta(*, failed_reason: Optional[str], reuse_domains: List[str]) -> Dict[str, Any]:
    return {
        "verification_status": "pending",
        "verification_report": None,
        "proof_status": "unanchored",
        "manifest_cid": None,
        "manifest_sha256": None,
        "manifest_gateway_url": None,
        "hcs_topic_id": None,
        "hcs_message_status": None,
        "anchor_payload": None,
        "anchored_at": None,
        "proof_last_error": None,
        "reuse_count": 0,
        "last_reused_at": None,
        "failed_reason": failed_reason,
        "reuse_domains": reuse_domains,
    }


def _coerce_meta(meta: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    merged = _default_meta(failed_reason=None, reuse_domains=[])
    if isinstance(meta, dict):
        merged.update(meta)
    if not isinstance(merged.get("reuse_domains"), list):
        merged["reuse_domains"] = []
    if not isinstance(merged.get("hol_sessions"), list):
        merged["hol_sessions"] = []
    merged["reuse_count"] = int(merged.get("reuse_count") or 0)
    return merged


def _tabular_summary(payload: bytes, delimiter: str) -> Tuple[bool, Dict[str, Any], Optional[str]]:
    try:
        text = payload.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = payload.decode("latin-1")
        except UnicodeDecodeError as exc:
            return False, {}, f"Unable to decode text payload: {exc}"

    row_count = 0
    max_columns = 0
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    for row in reader:
        row_count += 1
        max_columns = max(max_columns, len(row))
    return True, {"rows": row_count, "columns": max_columns}, None


def _json_summary(payload: bytes) -> Tuple[bool, Dict[str, Any], Optional[str]]:
    try:
        data = json.loads(payload.decode("utf-8"))
    except Exception as exc:  # noqa: BLE001
        return False, {}, f"Invalid JSON payload: {exc}"

    if isinstance(data, list):
        row_count = len(data)
        keys = set()
        for item in data:
            if isinstance(item, dict):
                keys.update(item.keys())
        return True, {"rows": row_count, "columns": len(keys), "shape": "list"}, None
    if isinstance(data, dict):
        return True, {"rows": 1, "columns": len(data.keys()), "shape": "object"}, None
    return True, {"rows": 1, "columns": 1, "shape": type(data).__name__}, None


def _xlsx_summary(payload: bytes) -> Tuple[bool, Dict[str, Any], Optional[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001
        return False, {}, f"openpyxl is unavailable for xlsx validation: {exc}"

    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook.active
        row_count = 0
        max_columns = 0
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_count += 1
                max_columns = max(max_columns, len(row))
        return True, {"rows": row_count, "columns": max_columns, "sheet": sheet.title}, None
    except Exception as exc:  # noqa: BLE001
        return False, {}, f"Invalid XLSX payload: {exc}"


def _zip_summary(payload: bytes) -> Tuple[bool, Dict[str, Any], Optional[str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
            files = [entry.filename for entry in archive.infolist() if not entry.is_dir()]
            return True, {"file_count": len(files), "sample_files": files[:10]}, None
    except Exception as exc:  # noqa: BLE001
        return False, {}, f"Invalid ZIP payload: {exc}"


def _txt_summary(payload: bytes) -> Tuple[bool, Dict[str, Any], Optional[str]]:
    try:
        text = payload.decode("utf-8")
    except UnicodeDecodeError:
        text = payload.decode("latin-1", errors="ignore")
    lines = text.splitlines()
    return True, {"rows": len(lines), "columns": 1}, None


def _extension_summary(extension: str, payload: bytes) -> Tuple[bool, Dict[str, Any], Optional[str]]:
    if extension == ".csv":
        return _tabular_summary(payload, ",")
    if extension == ".tsv":
        return _tabular_summary(payload, "\t")
    if extension == ".json":
        return _json_summary(payload)
    if extension == ".txt":
        return _txt_summary(payload)
    if extension == ".xlsx":
        return _xlsx_summary(payload)
    if extension == ".zip":
        return _zip_summary(payload)
    return False, {}, f"Unsupported extension {extension}"


def _build_verification_report(
    *,
    asset: DataAsset,
    payload: bytes,
    extension: str,
    db: Session,
) -> Tuple[str, Dict[str, Any]]:
    computed_hash = hashlib.sha256(payload).hexdigest()
    checksum_ok = computed_hash == asset.sha256
    not_empty = len(payload) > 0
    parse_ok, summary, parse_error = _extension_summary(extension, payload)

    duplicates = (
        db.query(DataAsset)
        .filter(DataAsset.sha256 == asset.sha256)
        .filter(DataAsset.id != asset.id)
        .all()
    )
    duplicate_ok = len(duplicates) == 0

    checks = {
        "checksum": {
            "passed": checksum_ok,
            "expected": asset.sha256,
            "computed": computed_hash,
        },
        "file_type_consistency": {
            "passed": parse_ok,
            "extension": extension,
            "error": parse_error,
        },
        "empty_file": {
            "passed": not_empty,
            "size_bytes": len(payload),
        },
        "duplicate_hash": {
            "passed": duplicate_ok,
            "duplicate_ids": [row.id for row in duplicates],
        },
    }
    status_value = "passed" if all(item["passed"] for item in checks.values()) else "failed"
    report = {
        "checked_at": _now_iso(),
        "status": status_value,
        "checks": checks,
        "summary": summary,
    }
    return status_value, report


def _serialize_asset(asset: DataAsset) -> Dict[str, Any]:
    meta = _coerce_meta(asset.meta)
    return {
        "id": asset.id,
        "title": asset.title,
        "description": asset.description,
        "lab_name": asset.lab_name,
        "uploader_name": asset.uploader_name,
        "data_classification": asset.data_classification,
        "tags": asset.tags or [],
        "intended_visibility": asset.intended_visibility or "private",
        "filename": asset.filename,
        "size_bytes": asset.size_bytes,
        "content_type": asset.content_type,
        "sha256": asset.sha256,
        "created_at": asset.created_at.isoformat() if asset.created_at else "",
        "verification_status": meta.get("verification_status") or "pending",
        "proof_status": meta.get("proof_status") or "unanchored",
        "manifest_cid": meta.get("manifest_cid"),
        "reuse_count": int(meta.get("reuse_count") or 0),
        "last_reused_at": meta.get("last_reused_at"),
        "failed_reason": meta.get("failed_reason"),
        "reuse_domains": meta.get("reuse_domains") or [],
    }


def _build_dataset_citation_payload(asset: DataAsset) -> Dict[str, Any]:
    meta = _coerce_meta(asset.meta)
    return {
        "type": "dataset",
        "title": asset.title,
        "dataset_id": asset.id,
        "authoring_lab": asset.lab_name,
        "uploader": asset.uploader_name,
        "published_at": asset.created_at.isoformat() if asset.created_at else _now_iso(),
        "provider": "Synaptica Data Agent",
        "classification": asset.data_classification,
        "failed_reason": meta.get("failed_reason"),
        "reuse_domains": meta.get("reuse_domains") or [],
        "identifiers": {
            "file_sha256": asset.sha256,
            "manifest_cid": meta.get("manifest_cid"),
            "manifest_sha256": meta.get("manifest_sha256"),
            "hcs_topic_id": meta.get("hcs_topic_id"),
        },
        "links": {
            "manifest_gateway_url": meta.get("manifest_gateway_url"),
            "proof_endpoint": f"/api/data-agent/datasets/{asset.id}/proof",
        },
    }


def _build_proof_bundle(asset: DataAsset) -> Dict[str, Any]:
    meta = _coerce_meta(asset.meta)
    return {
        "dataset_id": asset.id,
        "file_sha256": asset.sha256,
        "manifest_cid": meta.get("manifest_cid"),
        "manifest_sha256": meta.get("manifest_sha256"),
        "manifest_gateway_url": meta.get("manifest_gateway_url"),
        "hcs_topic_id": meta.get("hcs_topic_id"),
        "hcs_message_status": meta.get("hcs_message_status"),
        "anchor_payload": meta.get("anchor_payload"),
        "anchored_at": meta.get("anchored_at"),
        "verification_status": meta.get("verification_status") or "pending",
        "verification_report": meta.get("verification_report"),
        "proof_status": meta.get("proof_status") or "unanchored",
    }


def _hol_agent_summary(agent: HolAgentSummary) -> Dict[str, Any]:
    return {
        "uaid": getattr(agent, "uaid", ""),
        "name": getattr(agent, "name", ""),
        "description": getattr(agent, "description", ""),
        "capabilities": getattr(agent, "capabilities", []) or [],
        "categories": getattr(agent, "categories", []) or [],
        "transports": getattr(agent, "transports", []) or [],
        "pricing": getattr(agent, "pricing", {}) or {},
        "registry": getattr(agent, "registry", None),
        "available": getattr(agent, "available", None),
        "broker_marked_available": getattr(agent, "broker_marked_available", None),
        "availability_status": getattr(agent, "availability_status", None),
        "source_url": getattr(agent, "source_url", None),
        "adapter": getattr(agent, "adapter", None),
        "protocol": getattr(agent, "protocol", None),
        "synaptica_verified": bool(getattr(agent, "synaptica_verified", False)),
        "synaptica_verified_at": getattr(agent, "synaptica_verified_at", None),
        "synaptica_verification_mode": getattr(agent, "synaptica_verification_mode", None),
        "usability_tier": getattr(agent, "usability_tier", "exploratory"),
        "usability_reason": getattr(agent, "usability_reason", ""),
    }


def _matches_required_capabilities(
    agent: HolAgentSummary,
    required_capabilities: List[str],
) -> bool:
    if not required_capabilities:
        return True
    haystack = " ".join(
        [
            agent.name,
            agent.description,
            *agent.capabilities,
            *agent.categories,
            *agent.transports,
        ]
    ).lower()
    return any(capability.lower() in haystack for capability in required_capabilities if capability.strip())


def _hol_candidate_sort_key(agent: HolAgentSummary) -> tuple[int, int, int, int, int, int]:
    availability = str(getattr(agent, "availability_status", "") or "").strip().lower()
    source_url = getattr(agent, "source_url", None)
    transports = getattr(agent, "transports", []) or []
    tier = str(getattr(agent, "usability_tier", "exploratory") or "exploratory").strip().lower()
    tier_rank = {
        "verified": 3,
        "broker_available": 2,
        "exploratory": 1,
        "blocked": 0,
    }.get(tier, 0)
    has_url = 1 if source_url else 0
    has_http_transport = 1 if "http" in [str(item).lower() for item in transports] else 0
    available = 1 if getattr(agent, "broker_marked_available", None) is True else 0
    synaptica_verified = 1 if bool(getattr(agent, "synaptica_verified", False)) else 0
    online = 1 if availability in {"online", "ok", "available"} else 0
    stale = 1 if availability == "stale" else 0
    return (
        tier_rank,
        synaptica_verified,
        available,
        online,
        has_http_transport or has_url,
        0 if stale else 1,
    )


def _get_broker_chatable_rejection_reason(agent: HolAgentSummary) -> Optional[str]:
    transports = [str(item).strip().lower() for item in (getattr(agent, "transports", []) or []) if str(item).strip()]
    source_url = str(getattr(agent, "source_url", "") or "").strip()
    protocol = str(getattr(agent, "protocol", "") or "").strip().lower()
    adapter = str(getattr(agent, "adapter", "") or "").strip().lower()
    usability_tier = str(getattr(agent, "usability_tier", "exploratory") or "exploratory").strip().lower()
    usability_reason = str(getattr(agent, "usability_reason", "") or "").strip()

    if usability_tier == "blocked":
        return usability_reason or "Agent is blocked by current HOL usability signals."

    if "http" in transports:
        return None
    if protocol in {"a2a", "uagent"}:
        return f"Agent protocol {protocol} is not broker-chatable in this Synaptica flow."
    if adapter in {"a2a-registry-adapter", "agentverse-adapter"}:
        return f"Agent adapter {adapter} is not broker-chatable in this Synaptica flow."
    if source_url:
        return None
    return "Agent does not expose an HTTP transport or resolvable source URL."


def _is_broker_chatable_candidate(agent: HolAgentSummary) -> bool:
    return _get_broker_chatable_rejection_reason(agent) is None


def _build_dataset_hol_search_query(
    asset: DataAsset,
    request: DatasetHolUseRequest,
    required_capabilities: List[str],
) -> str:
    if request.search_query and request.search_query.strip():
        return request.search_query.strip()

    parts = [
        asset.title,
        asset.lab_name,
        asset.data_classification,
        "data agent",
        *list(asset.tags or []),
        *required_capabilities,
    ]
    return " ".join(str(part).strip() for part in parts if str(part).strip())


def _build_dataset_hol_search_queries(
    asset: DataAsset,
    request: DatasetHolUseRequest,
    required_capabilities: List[str],
) -> List[str]:
    fallbacks = [
        _build_dataset_hol_search_query(asset, request, required_capabilities),
        " ".join(["data agent", *required_capabilities]).strip(),
        "data agent",
    ]

    deduped: List[str] = []
    for item in fallbacks:
        normalized = " ".join(str(item or "").split())
        if normalized and normalized not in deduped:
            deduped.append(normalized)
    return deduped


def _build_dataset_hol_error_detail(
    *,
    message: str,
    search_queries: List[str],
    discovered: Optional[List[HolAgentSummary]] = None,
    rejected: Optional[List[Dict[str, Any]]] = None,
    attempted_errors: Optional[List[str]] = None,
) -> Dict[str, Any]:
    return DatasetHolUseErrorDetail(
        message=message,
        search_queries=search_queries,
        discovered_candidates=[_hol_agent_summary(agent) for agent in (discovered or [])][:10],
        rejected_candidates=(rejected or [])[:10],
        attempted_errors=(attempted_errors or [])[:10],
    ).model_dump()


def _build_dataset_hol_message(asset: DataAsset, instructions: Optional[str] = None) -> str:
    meta = _coerce_meta(asset.meta)
    proof_bundle = _build_proof_bundle(asset)
    citation = _build_dataset_citation_payload(asset)
    payload = {
        "type": "synaptica_dataset_task",
        "dataset": {
            "id": asset.id,
            "title": asset.title,
            "description": asset.description,
            "lab_name": asset.lab_name,
            "classification": asset.data_classification,
            "tags": asset.tags or [],
            "filename": asset.filename,
            "size_bytes": asset.size_bytes,
            "created_at": asset.created_at.isoformat() if asset.created_at else _now_iso(),
        },
        "verification": {
            "status": meta.get("verification_status"),
            "report": meta.get("verification_report"),
        },
        "proof": proof_bundle,
        "citation": citation,
        "instructions": instructions
        or (
            "Review this dataset context and provide a concise data-agent response focused on "
            "reuse potential, analysis suggestions, and any integrity or provenance concerns."
        ),
    }
    return json.dumps(payload, ensure_ascii=True, separators=(",", ":"))


def _append_hol_session_trace(
    asset: DataAsset,
    *,
    selected_agent: Dict[str, Any],
    session_id: str,
    search_query: Optional[str],
    instructions: Optional[str],
    transport: Optional[str],
    broker_response: Dict[str, Any],
) -> None:
    meta = _coerce_meta(asset.meta)
    sessions = list(meta.get("hol_sessions") or [])
    sessions.append(
        {
            "created_at": _now_iso(),
            "selected_agent": selected_agent,
            "session_id": session_id,
            "search_query": search_query,
            "instructions": instructions,
            "transport": transport,
            "broker_response": broker_response,
        }
    )
    meta["hol_sessions"] = sessions[-20:]
    asset.meta = meta


def _build_data_agent_chat_response(message: str, db: Session) -> str:
    query = (message or "").strip().lower()
    rows = db.query(DataAsset).order_by(DataAsset.created_at.desc()).limit(50).all()

    terms = re.findall(r"[a-z0-9]{3,}", query)
    if terms:
        matched: List[Tuple[int, DataAsset]] = []
        for row in rows:
            haystack = " ".join(
                [
                    row.title or "",
                    row.description or "",
                    row.lab_name or "",
                    row.data_classification or "",
                    " ".join(row.tags or []),
                ]
            ).lower()
            score = sum(1 for term in terms if term in haystack)
            if score:
                matched.append((score, row))
        matched.sort(key=lambda item: (item[0], item[1].created_at or datetime.min), reverse=True)
        rows = [row for _, row in matched[:5]]
    else:
        rows = rows[:5]

    if not rows:
        return (
            "No datasets are currently stored in the Synaptica Data Vault. "
            "Upload a failed or underused dataset first, then ask again."
        )

    lines = [
        "Synaptica Data Agent dataset summary:",
    ]
    for row in rows[:5]:
        meta = _coerce_meta(row.meta)
        lines.append(
            "- "
            f"{row.title} | lab={row.lab_name} | classification={row.data_classification} | "
            f"verification={meta.get('verification_status')} | proof={meta.get('proof_status')} | "
            f"tags={', '.join(row.tags or []) or 'none'}"
        )
    lines.append(
        "Reply with a dataset title, tag, lab, or classification if you want a narrower match."
    )
    return "\n".join(lines)


def _manifest_payload(asset: DataAsset) -> Dict[str, Any]:
    meta = _coerce_meta(asset.meta)
    return {
        "schema": "synaptica-data-asset-manifest-v1",
        "datasetId": asset.id,
        "title": asset.title,
        "description": asset.description,
        "labName": asset.lab_name,
        "uploaderName": asset.uploader_name,
        "dataClassification": asset.data_classification,
        "tags": asset.tags or [],
        "intendedVisibility": asset.intended_visibility,
        "filename": asset.filename,
        "sizeBytes": asset.size_bytes,
        "contentType": asset.content_type,
        "fileSha256": asset.sha256,
        "verificationStatus": meta.get("verification_status") or "pending",
        "verificationReport": meta.get("verification_report"),
        "failedReason": meta.get("failed_reason"),
        "reuseDomains": meta.get("reuse_domains") or [],
        "createdAt": asset.created_at.isoformat() if asset.created_at else None,
    }


def _pinata_headers() -> Dict[str, str]:
    jwt = os.getenv("PINATA_JWT")
    if jwt:
        return {
            "Authorization": f"Bearer {jwt}",
            "Content-Type": "application/json",
        }

    api_key = os.getenv("PINATA_API_KEY")
    secret_key = os.getenv("PINATA_SECRET_KEY")
    if not api_key or not secret_key:
        raise RuntimeError(
            "Pinata credentials missing. Configure PINATA_JWT or PINATA_API_KEY/PINATA_SECRET_KEY."
        )
    return {
        "pinata_api_key": api_key,
        "pinata_secret_api_key": secret_key,
        "Content-Type": "application/json",
    }


async def _pin_manifest_to_pinata(asset: DataAsset, manifest: Dict[str, Any]) -> Tuple[str, str]:
    headers = _pinata_headers()
    payload = {
        "pinataMetadata": {
            "name": f"data-asset-manifest-{asset.id}.json",
            "keyvalues": {
                "project": "Synaptica",
                "type": "data_asset_manifest",
                "dataset_id": asset.id,
            },
        },
        "pinataContent": manifest,
    }
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(PINATA_PIN_JSON_URL, headers=headers, json=payload)
        response.raise_for_status()
        body = response.json()
    except httpx.HTTPStatusError as exc:
        response_body = ""
        with contextlib.suppress(Exception):
            response_body = exc.response.text.strip()
        snippet = f" Body: {response_body[:240]}" if response_body else ""
        raise RuntimeError(
            f"Pinata request failed with status {exc.response.status_code}.{snippet}"
        ) from exc
    except httpx.HTTPError as exc:
        raise RuntimeError(f"Pinata request error: {exc}") from exc

    cid = body.get("IpfsHash")
    if not cid:
        raise RuntimeError("Pinata did not return IpfsHash")
    return cid, f"https://gateway.pinata.cloud/ipfs/{cid}"


async def _submit_anchor_message(anchor_payload: Dict[str, Any]) -> Tuple[str, str]:
    try:
        client = get_hedera_client()
    except ValidationError as exc:
        raise RuntimeError(
            "Hedera configuration missing. Configure HEDERA_ACCOUNT_ID and HEDERA_PRIVATE_KEY."
        ) from exc
    topic_id = client.topic_id
    if topic_id is None:
        topic_id = await client.create_topic("Synaptica Data Agent Provenance")
    message_status = await client.submit_message(
        json.dumps(anchor_payload, sort_keys=True, separators=(",", ":")),
        topic_id=topic_id,
    )
    return str(topic_id), str(message_status)


def _apply_verification(asset: DataAsset, db: Session) -> None:
    path = Path(asset.stored_path)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Stored file is missing")
    payload = path.read_bytes()
    extension = path.suffix.lower()
    verification_status, report = _build_verification_report(
        asset=asset,
        payload=payload,
        extension=extension,
        db=db,
    )
    meta = _coerce_meta(asset.meta)
    meta["verification_status"] = verification_status
    meta["verification_report"] = report
    asset.meta = meta


def _similar_datasets(db: Session, asset: DataAsset, limit: int = 5) -> List[Dict[str, Any]]:
    source_tags = set(tag.lower() for tag in (asset.tags or []))
    rows = (
        db.query(DataAsset)
        .filter(DataAsset.id != asset.id)
        .order_by(DataAsset.created_at.desc())
        .all()
    )

    scored: List[Tuple[float, DataAsset]] = []
    for row in rows:
        row_tags = set(tag.lower() for tag in (row.tags or []))
        overlap = len(source_tags.intersection(row_tags))
        score = float(overlap)
        if row.data_classification == asset.data_classification:
            score += 1.0
        if row.lab_name == asset.lab_name:
            score += 0.5
        if score > 0:
            scored.append((score, row))

    scored.sort(key=lambda item: item[0], reverse=True)
    output: List[Dict[str, Any]] = []
    for score, row in scored[:limit]:
        output.append(
            {
                "id": row.id,
                "title": row.title,
                "data_classification": row.data_classification,
                "tags": row.tags or [],
                "similarity_score": round(score, 2),
            }
        )
    return output


async def _anchor_dataset_internal(asset: DataAsset, db: Session) -> Dict[str, Any]:
    meta = _coerce_meta(asset.meta)

    manifest = _manifest_payload(asset)
    manifest_bytes = json.dumps(
        manifest,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    ).encode("utf-8")
    manifest_sha256 = hashlib.sha256(manifest_bytes).hexdigest()

    try:
        cid, gateway_url = await _pin_manifest_to_pinata(asset, manifest)
    except Exception as exc:  # noqa: BLE001
        meta["proof_status"] = "failed"
        meta["proof_last_error"] = f"Pinata upload failed: {exc}"
        asset.meta = meta
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to pin manifest to Pinata: {exc}",
        ) from exc

    anchor_payload = {
        "dataset_id": asset.id,
        "manifest_cid": cid,
        "manifest_sha256": manifest_sha256,
        "lab_name": asset.lab_name,
        "uploaded_at": asset.created_at.isoformat() if asset.created_at else _now_iso(),
    }

    try:
        topic_id, message_status = await _submit_anchor_message(anchor_payload)
    except Exception as exc:  # noqa: BLE001
        meta["proof_status"] = "manifest_pinned"
        meta["manifest_cid"] = cid
        meta["manifest_sha256"] = manifest_sha256
        meta["manifest_gateway_url"] = gateway_url
        meta["proof_last_error"] = f"Hedera anchoring failed: {exc}"
        asset.meta = meta
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to anchor manifest on Hedera: {exc}",
        ) from exc

    meta["proof_status"] = "anchored"
    meta["manifest_cid"] = cid
    meta["manifest_sha256"] = manifest_sha256
    meta["manifest_gateway_url"] = gateway_url
    meta["hcs_topic_id"] = topic_id
    meta["hcs_message_status"] = message_status
    meta["anchor_payload"] = anchor_payload
    meta["anchored_at"] = _now_iso()
    meta["proof_last_error"] = None
    asset.meta = meta
    db.commit()
    db.refresh(asset)
    return _build_proof_bundle(asset)


@router.post("/datasets", response_model=UploadDatasetResponse, status_code=status.HTTP_201_CREATED)
async def upload_dataset(
    title: str = Form(...),
    description: str = Form(""),
    lab_name: str = Form(...),
    data_classification: str = Form(...),
    tags: Optional[str] = Form(default=None),
    intended_visibility: str = Form("private"),
    uploader_name: Optional[str] = Form(default=None),
    failed_reason: Optional[str] = Form(default=None),
    reuse_domains: Optional[str] = Form(default=None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> UploadDatasetResponse:
    """Upload a dataset for Data Agent storage and run automated verification."""

    cleaned_title = title.strip()
    cleaned_lab = lab_name.strip()
    if not cleaned_title:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="title is required")
    if not cleaned_lab:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="lab_name is required")

    classification = _validate_classification(data_classification)
    visibility = _validate_visibility(intended_visibility)
    parsed_tags = _parse_string_list(tags, "tags")
    parsed_reuse_domains = _parse_string_list(reuse_domains, "reuse_domains")

    original_name = (file.filename or "").strip()
    if not original_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="file is required")

    extension = Path(original_name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type. Allowed extensions: {sorted(ALLOWED_EXTENSIONS)}",
        )

    payload = await file.read()
    size = len(payload)
    if size == 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")
    if size > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_CONTENT_TOO_LARGE,
            detail="File exceeds maximum size of 25MB",
        )

    safe_name = Path(original_name).name
    stored_name = f"{uuid.uuid4().hex}_{safe_name}"
    storage_dir = _get_storage_dir()
    storage_dir.mkdir(parents=True, exist_ok=True)
    stored_path = storage_dir / stored_name
    stored_path.write_bytes(payload)

    try:
        data_asset = DataAsset(  # type: ignore[call-arg]
            id=str(uuid.uuid4()),
            title=cleaned_title,
            description=description.strip() or None,
            lab_name=cleaned_lab,
            uploader_name=(uploader_name or "").strip() or None,
            data_classification=classification,
            tags=parsed_tags,
            intended_visibility=visibility,
            filename=safe_name,
            stored_path=str(stored_path),
            content_type=file.content_type,
            size_bytes=size,
            sha256=hashlib.sha256(payload).hexdigest(),
            meta=_default_meta(
                failed_reason=(failed_reason or "").strip() or None,
                reuse_domains=parsed_reuse_domains,
            ),
        )
        db.add(data_asset)
        db.flush()

        _apply_verification(data_asset, db)
        db.commit()
        db.refresh(data_asset)

        auto_anchor = os.getenv("DATA_AGENT_AUTO_ANCHOR", "0").lower() in {"1", "true", "yes"}
        if auto_anchor:
            try:
                await _anchor_dataset_internal(data_asset, db)
            except HTTPException:
                # Keep upload successful even when anchoring fails; proof_status captures failure.
                pass
            db.refresh(data_asset)
    except Exception:
        stored_path.unlink(missing_ok=True)
        raise

    return UploadDatasetResponse(**_serialize_asset(data_asset))


@router.post("/datasets/{dataset_id}/verify", response_model=DataAssetDetailResponse)
async def verify_dataset(dataset_id: str, db: Session = Depends(get_db)) -> DataAssetDetailResponse:
    """Manually trigger verification checks for an existing dataset."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    _apply_verification(asset, db)
    db.commit()
    db.refresh(asset)

    return DataAssetDetailResponse(
        **_serialize_asset(asset),
        verification_report=_coerce_meta(asset.meta).get("verification_report"),
        proof_bundle=_build_proof_bundle(asset),
        similar_datasets=[SimilarDataset(**row) for row in _similar_datasets(db, asset)],
        hol_sessions=_coerce_meta(asset.meta).get("hol_sessions") or [],
    )


@router.post("/datasets/{dataset_id}/anchor", response_model=DatasetProofResponse)
async def anchor_dataset(dataset_id: str, db: Session = Depends(get_db)) -> DatasetProofResponse:
    """Pin canonical manifest and anchor provenance payload to Hedera HCS."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    # Ensure trust report is fresh before anchoring.
    _apply_verification(asset, db)
    db.commit()
    db.refresh(asset)

    proof = await _anchor_dataset_internal(asset, db)
    return DatasetProofResponse(**proof)


@router.get("/datasets", response_model=DataAssetListResponse)
async def list_datasets(
    q: Optional[str] = Query(default=None),
    tag: Optional[str] = Query(default=None),
    classification: Optional[str] = Query(default=None),
    verification_status: Optional[str] = Query(default=None),
    proof_status: Optional[str] = Query(default=None),
    lab_name: Optional[str] = Query(default=None),
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
) -> DataAssetListResponse:
    """List datasets with discovery-focused filters."""

    query = db.query(DataAsset).order_by(DataAsset.created_at.desc())
    if classification:
        query = query.filter(DataAsset.data_classification == _validate_classification(classification))
    if lab_name:
        query = query.filter(DataAsset.lab_name.ilike(f"%{lab_name.strip()}%"))
    if q:
        term = f"%{q.strip()}%"
        query = query.filter(
            DataAsset.title.ilike(term)
            | DataAsset.description.ilike(term)
            | DataAsset.lab_name.ilike(term)
        )

    records = query.all()
    if tag:
        tag_lower = tag.strip().lower()
        records = [
            row
            for row in records
            if any((asset_tag or "").strip().lower() == tag_lower for asset_tag in (row.tags or []))
        ]
    if verification_status:
        normalized = verification_status.strip().lower()
        records = [
            row for row in records if (_coerce_meta(row.meta).get("verification_status") or "").lower() == normalized
        ]
    if proof_status:
        normalized = proof_status.strip().lower()
        records = [row for row in records if (_coerce_meta(row.meta).get("proof_status") or "").lower() == normalized]

    total = len(records)
    paged = records[offset : offset + limit]
    return DataAssetListResponse(
        total=total,
        limit=limit,
        offset=offset,
        datasets=[DataAssetResponse(**_serialize_asset(row)) for row in paged],
    )


@router.get("/datasets/{dataset_id}", response_model=DataAssetDetailResponse)
async def get_dataset(dataset_id: str, db: Session = Depends(get_db)) -> DataAssetDetailResponse:
    """Fetch detailed dataset metadata with trust and similarity context."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    meta = _coerce_meta(asset.meta)
    return DataAssetDetailResponse(
        **_serialize_asset(asset),
        verification_report=meta.get("verification_report"),
        proof_bundle=_build_proof_bundle(asset),
        similar_datasets=[SimilarDataset(**row) for row in _similar_datasets(db, asset)],
        hol_sessions=meta.get("hol_sessions") or [],
    )


@router.post("/datasets/{dataset_id}/hol-use", response_model=DatasetHolUseResponse)
async def use_hol_data_agent(
    dataset_id: str,
    request: DatasetHolUseRequest,
    db: Session = Depends(get_db),
) -> DatasetHolUseResponse:
    """Use a HOL data agent against a dataset via broker chat session/message."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    try:
        await run_in_threadpool(hol_check_sidecar_health)
    except HolClientError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc)) from exc

    required_capabilities = [
        item.strip()
        for item in (request.required_capabilities or DEFAULT_HOL_DATA_CAPABILITIES)
        if isinstance(item, str) and item.strip()
    ]
    search_queries = _build_dataset_hol_search_queries(asset, request, required_capabilities)
    search_query = search_queries[0]

    if request.uaid and request.uaid.strip():
        explicit_summary = HolAgentSummary(
            uaid=request.uaid.strip(),
            name=request.uaid.strip(),
            description="",
            capabilities=[],
            categories=[],
            transports=[],
            pricing={},
            registry=None,
            available=None,
            broker_marked_available=None,
            availability_status=None,
            source_url=None,
            adapter=None,
            protocol=None,
        )
        verification_map = get_hol_agent_verification_map(db, [explicit_summary.uaid])
        apply_hol_agent_usability(explicit_summary, verification_map.get(explicit_summary.uaid))
        selected_agent = _hol_agent_summary(explicit_summary)
        candidate_agents: List[HolAgentSummary] = []
    else:
        discovered: List[HolAgentSummary] = []
        seen_uaids: set[str] = set()
        search_errors: List[str] = []
        broker_limit = min(100, max(request.limit, request.limit * 5))
        for candidate_query in search_queries:
            try:
                results = await run_in_threadpool(
                    hol_search_agents,
                    candidate_query,
                    limit=broker_limit,
                )
            except HolClientError as exc:
                search_errors.append(str(exc))
                continue

            for agent in results:
                uaid = str(getattr(agent, "uaid", "") or "").strip()
                if not uaid or uaid in seen_uaids:
                    continue
                seen_uaids.add(uaid)
                discovered.append(agent)

        verification_map = get_hol_agent_verification_map(db, [agent.uaid for agent in discovered])
        for agent in discovered:
            apply_hol_agent_usability(agent, verification_map.get(agent.uaid))

        if not discovered and search_errors:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=_build_dataset_hol_error_detail(
                    message="HOL discovery failed while searching for data agents.",
                    search_queries=search_queries,
                    attempted_errors=search_errors,
                ),
            )

        candidates = [
            agent for agent in discovered if _matches_required_capabilities(agent, required_capabilities)
        ]
        if not candidates:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=_build_dataset_hol_error_detail(
                    message="No HOL agents found for the provided dataset query.",
                    search_queries=search_queries,
                    discovered=discovered,
                ),
            )

        rejected_candidates: List[Dict[str, Any]] = []
        broker_chatable_candidates: List[HolAgentSummary] = []
        for agent in candidates:
            rejection_reason = _get_broker_chatable_rejection_reason(agent)
            if rejection_reason:
                rejected_candidates.append(
                    {
                        **_hol_agent_summary(agent),
                        "reason": rejection_reason,
                    }
                )
                continue
            broker_chatable_candidates.append(agent)
        if not broker_chatable_candidates:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=_build_dataset_hol_error_detail(
                    message="No broker-chatable HOL agents found for the provided dataset query.",
                    search_queries=search_queries,
                    discovered=candidates,
                    rejected=rejected_candidates,
                ),
            )

        candidate_agents = sorted(
            broker_chatable_candidates,
            key=_hol_candidate_sort_key,
            reverse=True,
        )
        selected_agent = _hol_agent_summary(candidate_agents[0])

    message = _build_dataset_hol_message(asset, request.instructions)
    attempted_errors: List[str] = []
    if request.uaid and request.uaid.strip():
        candidates_to_try = [selected_agent]
    else:
        candidates_to_try = [_hol_agent_summary(agent) for agent in candidate_agents]

    session_id: Optional[str] = None
    broker_response: Optional[Dict[str, Any]] = None
    for candidate in candidates_to_try:
        try:
            session_id = await run_in_threadpool(
                hol_create_session,
                candidate["uaid"],
                transport=request.transport,
                as_uaid=request.as_uaid,
            )
        except HolClientError as exc:
            attempted_errors.append(f"{candidate['uaid']}: {exc}")
            if is_hol_hard_availability_failure(str(exc)):
                record_hol_agent_hard_failure(
                    db,
                    candidate["uaid"],
                    reason=str(exc),
                    transport=request.transport,
                )
            if should_use_direct_chat_fallback(exc):
                fallback_reason = str(exc)
                try:
                    broker_response = await run_in_threadpool(
                        hol_send_message,
                        None,
                        message,
                        uaid=candidate["uaid"],
                        as_uaid=request.as_uaid,
                    )
                    session_id = f"{HOL_DIRECT_SESSION_PREFIX}{uuid.uuid4()}"
                    broker_response = coerce_hol_broker_response(
                        broker_response,
                        mode="direct",
                        fallback_reason=fallback_reason,
                    )
                    selected_agent = candidate
                    break
                except HolClientError as direct_exc:
                    attempted_errors.append(f"{candidate['uaid']} (direct): {direct_exc}")
                    if is_hol_hard_availability_failure(str(direct_exc)):
                        record_hol_agent_hard_failure(
                            db,
                            candidate["uaid"],
                            reason=str(direct_exc),
                            transport=request.transport,
                        )
            continue
        try:
            broker_response = await run_in_threadpool(
                hol_send_message,
                session_id,
                message,
                as_uaid=request.as_uaid,
            )
            broker_response = coerce_hol_broker_response(
                broker_response,
                mode="session",
            )
            selected_agent = candidate
            break
        except HolClientError as exc:
            attempted_errors.append(f"{candidate['uaid']}: {exc}")
            if is_hol_hard_availability_failure(str(exc)):
                record_hol_agent_hard_failure(
                    db,
                    candidate["uaid"],
                    reason=str(exc),
                    transport=request.transport,
                )
            continue

    if session_id is None or broker_response is None:
        detail = attempted_errors[0] if len(attempted_errors) == 1 else "; ".join(attempted_errors[:5])
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=_build_dataset_hol_error_detail(
                message=detail,
                search_queries=search_queries,
                rejected=[selected_agent] if request.uaid and request.uaid.strip() else None,
                attempted_errors=attempted_errors,
            ),
        )

    success_mode = str(broker_response.get("mode") or "session").strip() or "session"
    verification_row = record_hol_agent_success(
        db,
        selected_agent["uaid"],
        mode=success_mode,
        transport=request.transport,
    )
    apply_hol_agent_usability(selected_agent, verification_row)
    _append_hol_session_trace(
        asset,
        selected_agent=selected_agent,
        session_id=session_id,
        search_query=None if request.uaid else search_query,
        instructions=request.instructions,
        transport=request.transport,
        broker_response=broker_response,
    )
    db.commit()
    db.refresh(asset)

    return DatasetHolUseResponse(
        success=True,
        selected_agent=selected_agent,
        session_id=session_id,
        broker_response=broker_response,
    )


@router.get("/datasets/{dataset_id}/proof", response_model=DatasetProofResponse)
async def get_dataset_proof(dataset_id: str, db: Session = Depends(get_db)) -> DatasetProofResponse:
    """Return proof bundle used for judge-facing verifiability demo."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")
    return DatasetProofResponse(**_build_proof_bundle(asset))


@router.get("/datasets/{dataset_id}/citation", response_model=DatasetCitationResponse)
async def get_dataset_citation(dataset_id: str, db: Session = Depends(get_db)) -> DatasetCitationResponse:
    """Return a JSON citation object for copy/paste into reports and tools."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")
    return DatasetCitationResponse(citation=_build_dataset_citation_payload(asset))


@router.post("/datasets/{dataset_id}/reuse-events", response_model=ReuseEventResponse)
async def record_dataset_reuse(dataset_id: str, db: Session = Depends(get_db)) -> ReuseEventResponse:
    """Increment reuse counters for a dataset."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    meta = _coerce_meta(asset.meta)
    meta["reuse_count"] = int(meta.get("reuse_count") or 0) + 1
    meta["last_reused_at"] = _now_iso()
    asset.meta = meta
    db.commit()
    db.refresh(asset)

    return ReuseEventResponse(
        dataset_id=asset.id,
        reuse_count=int(meta["reuse_count"]),
        last_reused_at=meta["last_reused_at"],
        message="Reuse event recorded.",
    )


@router.get("/datasets/{dataset_id}/download")
async def download_dataset(dataset_id: str, db: Session = Depends(get_db)) -> FileResponse:
    """Download the original dataset file."""

    asset = db.query(DataAsset).filter(DataAsset.id == dataset_id).one_or_none()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dataset not found")

    path = Path(asset.stored_path)
    if not path.exists() or not path.is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Stored file is missing for this dataset",
        )

    return FileResponse(
        path=str(path),
        media_type=asset.content_type or "application/octet-stream",
        filename=asset.filename,
    )


def _build_data_agent_card() -> AgentCard:
    return AgentCard(
        id="data-agent-001",
        name="Data Agent",
        description="Synaptica Data Vault agent for cataloging and summarizing stored datasets.",
        version="0.1.0",
        capabilities=[
            AgentCapability(name="dataset-search", description="Find datasets by title, lab, tag, or classification."),
            AgentCapability(name="dataset-summary", description="Summarize stored datasets and provenance state."),
        ],
        tags=["data", "desci", "a2a", "synaptica"],
        extras={"message_endpoint": "/api/data-agent/agent/a2a/v1/messages"},
    )


@router.get("/agent/.well-known/agent-card.json")
@router.get("/agent/.well-known/agent.json")
async def data_agent_card(request: Request) -> Dict[str, Any]:
    """Expose the built-in Data Agent as a broker-friendly A2A-compatible card."""

    agent_card = _build_data_agent_card()
    rpc_url = str(request.url_for("data_agent_rpc"))
    return build_agent_card_payload(agent_card, rpc_url=rpc_url)


@router.get("/agent/health")
async def data_agent_health() -> Dict[str, Any]:
    """Health endpoint for the built-in Data Agent A2A surface."""

    return {
        "status": "ok",
        "agent_id": "data-agent-001",
        "service": "synaptica-data-agent",
    }


@router.post("/agent", name="data_agent_rpc")
async def data_agent_rpc(
    payload: Dict[str, Any],
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Expose the built-in Data Agent over a minimal JSON-RPC A2A surface."""

    rpc_id = payload.get("id") if isinstance(payload, dict) else None
    method = str(payload.get("method") or "").strip() if isinstance(payload, dict) else ""
    if method not in {"message/send", "tasks/send"}:
        return build_error_response(
            rpc_id=rpc_id,
            code=-32601,
            message="Unsupported A2A method",
        )

    message, metadata = extract_message_text_and_metadata(payload)
    if not message:
        return build_error_response(
            rpc_id=rpc_id,
            code=-32602,
            message="A2A message payload must include at least one text part",
        )

    response = _build_data_agent_chat_response(message, db)
    task_id = None
    params = payload.get("params")
    if isinstance(params, dict):
        raw_task_id = params.get("id")
        if isinstance(raw_task_id, str) and raw_task_id.strip():
            task_id = raw_task_id.strip()
    return build_completed_task_response(
        rpc_id=rpc_id,
        text=response,
        task_id=task_id,
        metadata=metadata,
    )


@router.post("/agent/a2a/v1/messages", response_model=MessageResponse)
async def data_agent_message(
    payload: MessagePayload,
    db: Session = Depends(get_db),
) -> MessageResponse:
    """Respond to simple broker/A2A-style chat messages for the built-in Data Agent."""

    response = _build_data_agent_chat_response(payload.message, db)
    return MessageResponse(
        message_id=uuid.uuid4().hex,
        response=response,
        metadata=payload.metadata,
    )
